"""
Admin routes for user and attendance management.
Admin-only endpoints for managing users, attendance, and reports.
"""

from flask import Blueprint, request, g, send_file
from datetime import datetime, timedelta
import pytz
import io

from models.user import User
from models.attendance import Attendance
from models.face_embedding import FaceEmbedding
from middleware.auth_middleware import token_required, admin_required
from utils.response import success_response, error_response, paginated_response

# Timezone WIB
WIB = pytz.timezone('Asia/Jakarta')

admin_bp = Blueprint('admin', __name__, url_prefix='/api/admin')


def init_admin_routes(db):
    """
    Initialize admin routes with database connection.
    
    Args:
        db: MongoDB database instance
    """
    user_model = User(db)
    attendance_model = Attendance(db)
    face_embedding_model = FaceEmbedding(db)
    
    @admin_bp.route('/users', methods=['GET'])
    @token_required
    @admin_required
    def get_users():
        """
        Get all users with pagination and filters.
        
        Query Parameters:
            role: Filter by role (student/teacher/admin)
            class: Filter by class_name
            search: Search by name or NIS
            page: Page number (default: 1)
            per_page: Items per page (default: 20)
        
        Returns:
            200: Users list with pagination
        """
        try:
            # Parse query parameters
            role = request.args.get('role')
            class_name = request.args.get('class')
            search = request.args.get('search')
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('per_page', 20))
            
            # Build query
            query = {'deleted': {'$ne': True}}
            
            if role:
                query['role'] = role
            
            if class_name:
                query['class_name'] = class_name
            
            if search:
                # Search by name or NIS
                query['$or'] = [
                    {'name': {'$regex': search, '$options': 'i'}},
                    {'full_name': {'$regex': search, '$options': 'i'}},
                    {'nis': {'$regex': search, '$options': 'i'}}
                ]
            
            # Get total count
            total = user_model.collection.count_documents(query)
            
            # Calculate pagination
            skip = (page - 1) * per_page
            
            # Get users
            users = list(user_model.collection.find(query)
                        .skip(skip)
                        .limit(per_page)
                        .sort('created_at', -1))
            
            # Sanitize users
            for user in users:
                user['_id'] = str(user['_id'])
                user.pop('password_hash', None)
            
            return paginated_response(users, page, per_page, total, 'Users retrieved successfully')
            
        except Exception as e:
            return error_response(f'Failed to get users: {str(e)}', 500)
    
    @admin_bp.route('/users/<user_id>', methods=['DELETE'])
    @token_required
    @admin_required
    def delete_user(user_id):
        """
        Soft delete user and related face embeddings.
        
        Args:
            user_id: User ID to delete
        
        Returns:
            200: User deleted successfully
        """
        try:
            # Check if user exists
            user = user_model.find_by_user_id(user_id)
            if not user:
                return error_response('User not found', 404)
            
            # Soft delete user
            success = user_model.delete_user(user_id)
            
            if not success:
                return error_response('Failed to delete user', 500)
            
            # Delete face embeddings
            face_embedding_model.delete_by_user_id(user_id)
            
            return success_response(None, 'User deleted successfully')
            
        except Exception as e:
            return error_response(f'Failed to delete user: {str(e)}', 500)
    
    @admin_bp.route('/users/<user_id>', methods=['PUT'])
    @token_required
    @admin_required
    def update_user(user_id):
        """
        Update user data.
        
        Request Body:
            {
                "name": "New Name",
                "class_name": "XII-IPA-1",
                "phone": "08123456789",
                "nis": "2024001"
            }
        
        Returns:
            200: User updated successfully
        """
        try:
            data = request.get_json()
            
            # Allowed fields to update
            allowed_fields = ['name', 'full_name', 'class_name', 'phone', 'nis', 'role']
            updates = {k: v for k, v in data.items() if k in allowed_fields}
            
            if not updates:
                return error_response('No valid fields to update', 400)
            
            # Update user
            updated_user = user_model.update_user(user_id, updates)
            
            if not updated_user:
                return error_response('User not found', 404)
            
            return success_response(updated_user, 'User updated successfully')
            
        except Exception as e:
            return error_response(f'Failed to update user: {str(e)}', 500)
    
    @admin_bp.route('/attendance/daily', methods=['GET'])
    @token_required
    @admin_required
    def get_daily_attendance():
        """
        Get daily attendance for all students.
        
        Query Parameters:
            date: Date (YYYY-MM-DD) (default: today)
            class: Filter by class_name
        
        Returns:
            200: Daily attendance with statistics
        """
        try:
            # Parse parameters
            date_str = request.args.get('date')
            class_name = request.args.get('class')
            
            if date_str:
                date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                date = datetime.now(WIB)
            
            # Get start and end of day in WIB
            start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
            if not start_of_day.tzinfo:
                start_of_day = WIB.localize(start_of_day)
            end_of_day = start_of_day + timedelta(days=1)
            
            # Convert to UTC
            start_utc = start_of_day.astimezone(pytz.UTC).replace(tzinfo=None)
            end_utc = end_of_day.astimezone(pytz.UTC).replace(tzinfo=None)
            
            # Get attendance records
            query = {
                'check_in_time': {
                    '$gte': start_utc,
                    '$lt': end_utc
                }
            }
            
            attendance_records = list(attendance_model.collection.find(query))
            
            # Get all students
            user_query = {'role': 'student', 'deleted': {'$ne': True}}
            if class_name:
                user_query['class_name'] = class_name
            
            all_students = list(user_model.collection.find(user_query))
            
            # Build attendance map
            attendance_map = {}
            for record in attendance_records:
                attendance_map[record['user_id']] = record
            
            # Build result
            result = []
            present_count = 0
            late_count = 0
            absent_count = 0
            
            for student in all_students:
                user_id = student['user_id']
                attendance = attendance_map.get(user_id)
                
                if attendance:
                    status = attendance['status']
                    if status == 'present':
                        present_count += 1
                    elif status == 'late':
                        late_count += 1
                    
                    # Return UTC timestamp directly
                    check_in_utc = attendance['check_in_time'].replace(tzinfo=pytz.UTC)
                    
                    result.append({
                        'user_id': user_id,
                        'name': student.get('full_name') or student.get('name'),
                        'nis': student.get('nis'),
                        'class': student.get('class_name'),
                        'status': status,
                        'check_in_time': check_in_utc.isoformat(),
                        'confidence': attendance.get('confidence_score')
                    })
                else:
                    absent_count += 1
                    result.append({
                        'user_id': user_id,
                        'name': student.get('full_name') or student.get('name'),
                        'nis': student.get('nis'),
                        'class': student.get('class_name'),
                        'status': 'absent',
                        'check_in_time': None,
                        'confidence': None
                    })
            
            return success_response({
                'date': date.strftime('%Y-%m-%d'),
                'statistics': {
                    'total': len(all_students),
                    'present': present_count,
                    'late': late_count,
                    'absent': absent_count
                },
                'records': result
            }, 'Daily attendance retrieved successfully')
            
        except Exception as e:
            return error_response(f'Failed to get daily attendance: {str(e)}', 500)
    
    @admin_bp.route('/attendance/report', methods=['GET'])
    @token_required
    @admin_required
    def get_attendance_report():
        """
        Get attendance report for a period.
        
        Query Parameters:
            month: Month (1-12)
            year: Year
            class: Filter by class_name
        
        Returns:
            200: Attendance report with statistics
        """
        try:
            month = int(request.args.get('month', datetime.now().month))
            year = int(request.args.get('year', datetime.now().year))
            class_name = request.args.get('class')
            
            # Get first and last day of month
            first_day = datetime(year, month, 1, tzinfo=WIB)
            if month == 12:
                last_day = datetime(year + 1, 1, 1, tzinfo=WIB)
            else:
                last_day = datetime(year, month + 1, 1, tzinfo=WIB)
            
            # Convert to UTC
            first_day_utc = first_day.astimezone(pytz.UTC).replace(tzinfo=None)
            last_day_utc = last_day.astimezone(pytz.UTC).replace(tzinfo=None)
            
            # Get all students
            user_query = {'role': 'student', 'deleted': {'$ne': True}}
            if class_name:
                user_query['class_name'] = class_name
            
            students = list(user_model.collection.find(user_query))
            
            # Get attendance for period
            attendance_query = {
                'check_in_time': {
                    '$gte': first_day_utc,
                    '$lt': last_day_utc
                }
            }
            
            attendance_records = list(attendance_model.collection.find(attendance_query))
            
            # Build report per student
            report = []
            for student in students:
                user_id = student['user_id']
                
                # Filter attendance for this student
                user_attendance = [r for r in attendance_records if r['user_id'] == user_id]
                
                present = sum(1 for r in user_attendance if r['status'] == 'present')
                late = sum(1 for r in user_attendance if r['status'] == 'late')
                
                # Calculate working days in month (simplified: assume 20 working days)
                working_days = 20
                absent = working_days - (present + late)
                
                percentage = ((present + late) / working_days * 100) if working_days > 0 else 0
                
                report.append({
                    'user_id': user_id,
                    'name': student.get('full_name') or student.get('name'),
                    'nis': student.get('nis'),
                    'class': student.get('class_name'),
                    'present': present,
                    'late': late,
                    'absent': absent,
                    'percentage': round(percentage, 2)
                })
            
            return success_response({
                'period': {
                    'month': month,
                    'year': year
                },
                'report': report
            }, 'Report generated successfully')
            
        except Exception as e:
            return error_response(f'Failed to generate report: {str(e)}', 500)
    
    @admin_bp.route('/export/excel', methods=['POST'])
    @token_required
    @admin_required
    def export_excel():
        """
        Export attendance data to Excel.
        
        Request Body:
            {
                "start_date": "2024-12-01",
                "end_date": "2024-12-31",
                "class": "XII-IPA-1"
            }
        
        Returns:
            200: Excel file download
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            data = request.get_json()
            start_date_str = data.get('start_date')
            end_date_str = data.get('end_date')
            class_name = data.get('class')
            
            if not start_date_str or not end_date_str:
                return error_response('start_date and end_date are required', 400)
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Header
            headers = ['No', 'NIS', 'Name', 'Class', 'Date', 'Check-in Time', 'Status', 'Confidence']
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Get data
            start_utc = WIB.localize(start_date).astimezone(pytz.UTC).replace(tzinfo=None)
            end_utc = WIB.localize(end_date + timedelta(days=1)).astimezone(pytz.UTC).replace(tzinfo=None)
            
            query = {
                'check_in_time': {
                    '$gte': start_utc,
                    '$lt': end_utc
                }
            }
            
            records = list(attendance_model.collection.find(query).sort('check_in_time', -1))
            
            # Add data rows
            row_num = 2
            for idx, record in enumerate(records, 1):
                user = user_model.find_by_user_id(record['user_id'])
                
                if class_name and user.get('class_name') != class_name:
                    continue
                
                check_in_utc = record['check_in_time'].replace(tzinfo=pytz.UTC)
                check_in_wib = check_in_utc.astimezone(WIB)
                
                ws.append([
                    idx,
                    user.get('nis', '-'),
                    user.get('full_name') or user.get('name', '-'),
                    user.get('class_name', '-'),
                    check_in_wib.strftime('%Y-%m-%d'),
                    check_in_wib.strftime('%H:%M:%S'),
                    record['status'].upper(),
                    f"{record.get('confidence_score', 0):.2f}" if record.get('confidence_score') else '-'
                ])
                row_num += 1
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'attendance_report_{start_date_str}_to_{end_date_str}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            return error_response('openpyxl library not installed. Run: pip install openpyxl', 500)
        except Exception as e:
            return error_response(f'Failed to export Excel: {str(e)}', 500)
    
    return admin_bp
